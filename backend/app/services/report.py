"""Агрегация статистики игры и экспорт в CSV/Excel (PLAN §10)."""
from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Answer, Player, Question, Room
from app.models.enums import AnswerStatus


def _ranking(players: list[Player]) -> list[dict[str, Any]]:
    ordered = sorted(players, key=lambda p: p.score, reverse=True)
    out, last, place = [], None, 0
    for i, p in enumerate(ordered):
        if p.score != last:
            place, last = i + 1, p.score
        out.append({
            "place": place, "player_id": p.id,
            "name": f"{p.name} {p.surname}", "group_name": p.group_name,
            "score": p.score,
        })
    return out


_CORRECT = {AnswerStatus.AUTO_CORRECT, AnswerStatus.APPROVED}
_WRONG = {AnswerStatus.AUTO_WRONG, AnswerStatus.REJECTED, AnswerStatus.TIMEOUT}


async def build_report(session: AsyncSession, room: Room) -> dict[str, Any]:
    players = list((await session.execute(
        select(Player).where(Player.room_id == room.id)
    )).scalars().all())
    answers = list((await session.execute(
        select(Answer).where(Answer.room_id == room.id)
    )).scalars().all())

    by_player: dict[int, dict[str, int]] = {
        p.id: {"correct": 0, "wrong": 0, "pending": 0, "total": 0} for p in players
    }
    by_question: dict[int, dict[str, int]] = {}
    for a in answers:
        ps = by_player.get(a.player_id)
        if ps is not None:
            ps["total"] += 1
            if a.status in _CORRECT:
                ps["correct"] += 1
            elif a.status in _WRONG:
                ps["wrong"] += 1
            else:
                ps["pending"] += 1
        if a.question_id is not None:
            qs = by_question.setdefault(
                a.question_id, {"correct": 0, "wrong": 0, "total": 0}
            )
            qs["total"] += 1
            if a.status in _CORRECT:
                qs["correct"] += 1
            elif a.status in _WRONG:
                qs["wrong"] += 1

    player_rows = []
    name_by_id = {p.id: f"{p.name} {p.surname}" for p in players}
    group_by_id = {p.id: p.group_name for p in players}
    for p in players:
        st = by_player[p.id]
        answered = st["correct"] + st["wrong"]
        success = round(100 * st["correct"] / answered, 1) if answered else 0.0
        player_rows.append({
            "player_id": p.id, "name": name_by_id[p.id], "group_name": group_by_id[p.id],
            "score": p.score, "correct": st["correct"], "wrong": st["wrong"],
            "pending": st["pending"], "success_rate": success,
            "eliminated": p.eliminated,
        })

    # Тексты вопросов
    q_ids = list(by_question.keys())
    texts: dict[int, str] = {}
    if q_ids:
        for q in (await session.execute(
            select(Question).where(Question.id.in_(q_ids))
        )).scalars().all():
            texts[q.id] = q.text
    question_rows = [
        {
            "question_id": qid, "text": texts.get(qid, ""),
            "correct": st["correct"], "wrong": st["wrong"], "total": st["total"],
            "success_rate": round(100 * st["correct"] / st["total"], 1) if st["total"] else 0.0,
        }
        for qid, st in by_question.items()
    ]

    return {
        "room": {"id": room.id, "code": room.code, "subject": room.subject,
                 "status": room.status.value},
        "ranking": _ranking(players),
        "players": player_rows,
        "questions": question_rows,
    }


def report_to_csv(report: dict[str, Any]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Рейтинг"])
    w.writerow(["Место", "Игрок", "Группа", "Баллы"])
    for r in report["ranking"]:
        w.writerow([r["place"], r["name"], r["group_name"], r["score"]])
    w.writerow([])
    w.writerow(["Статистика по игрокам"])
    w.writerow(["Игрок", "Группа", "Баллы", "Верно", "Неверно", "На проверке", "% успеха"])
    for p in report["players"]:
        w.writerow([p["name"], p["group_name"], p["score"], p["correct"],
                    p["wrong"], p["pending"], p["success_rate"]])
    w.writerow([])
    w.writerow(["Статистика по вопросам"])
    w.writerow(["Вопрос", "Верно", "Неверно", "Всего", "% успеха"])
    for q in report["questions"]:
        w.writerow([q["text"], q["correct"], q["wrong"], q["total"], q["success_rate"]])
    return buf.getvalue().encode("utf-8-sig")


def report_to_xlsx(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Рейтинг"
    ws1.append(["Место", "Игрок", "Группа", "Баллы"])
    for r in report["ranking"]:
        ws1.append([r["place"], r["name"], r["group_name"], r["score"]])

    ws2 = wb.create_sheet("Игроки")
    ws2.append(["Игрок", "Группа", "Баллы", "Верно", "Неверно", "На проверке", "% успеха"])
    for p in report["players"]:
        ws2.append([p["name"], p["group_name"], p["score"], p["correct"],
                    p["wrong"], p["pending"], p["success_rate"]])

    ws3 = wb.create_sheet("Вопросы")
    ws3.append(["Вопрос", "Верно", "Неверно", "Всего", "% успеха"])
    for q in report["questions"]:
        ws3.append([q["text"], q["correct"], q["wrong"], q["total"], q["success_rate"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
